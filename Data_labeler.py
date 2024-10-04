from docx import Document
import re
import os
from openpyxl import Workbook, load_workbook

# Define the PII detection patterns
patterns_ = {
    'Aadhaar Number': [
        r'\b\d{12}\b',  # 12 consecutive digits
        r'\b\d{4} \d{4} \d{4}\b',  # 4-4-4 format with spaces
        r'\b\d{4}-\d{4}-\d{4}\b',  # 4-4-4 format with dashes
        r'\b(?:Aadhaar\s*No\.?|UID)\s*[:\-]?\s*\d{4}[ -]?\d{4}[ -]?\d{4}\b'  # Aadhaar No
    ],
    'Email Address': [
        r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    ],
    'Credit Card Number': [
        r'\b(?:\d[ -]*?){13,16}\b'  # 13 to 16 digits with  spaces or dashes
    ]
}


def extract_text_from_docx(file_path):
    try:
        doc = Document(file_path)
        text = [paragraph.text for paragraph in doc.paragraphs if paragraph.text]
        return '\n'.join(text)
    except Exception as e:
        print(f"Error reading Word document: {e}")
        return ""


def detect_pii(text):
    pii_found = {}

    
    if text.startswith("+") and re.search(r'\+\d{12}\b', text):
        pii_found['Phone Number'] = [text]  
        return pii_found  
    

    aadhaar_matches = re.findall('|'.join(patterns_['Aadhaar Number']), text)
    if aadhaar_matches:
        pii_found['Aadhaar Number'] = aadhaar_matches

    
    email_matches = re.findall(patterns_['Email Address'][0], text)
    if email_matches:
        pii_found['Email Address'] = email_matches

    
    cc_matches = re.findall(patterns_['Credit Card Number'][0], text)
    if cc_matches:
        pii_found['Credit Card Number'] = cc_matches

    return pii_found  


def label_and_append_data(file_name):
    
    file_path = fr"C:\Users\Lokghesh VAK\OneDrive\Desktop\TEST_FILES\{file_name}.docx"
    extracted_text = extract_text_from_docx(file_path)
    
    if not extracted_text:
        print("No text extracted from document. Please check the document path and content.")
        return
    
  
    data = [{"text": line.strip()} for line in extracted_text.split('\n') if line.strip()]
    
    if not data:
        print("No lines of text found for processing.")
        return
    
    
    for entry in data:
        pii_types = detect_pii(entry["text"])
        entry["label"] = "PII" if pii_types else "non-PII"
        entry["type_of_pii"] = ', '.join(pii_types.keys()) if pii_types else "None"

    
    excel_file_path = r'C:\Users\Lokghesh VAK\OneDrive\Desktop\TEST_FILES\Data_Mana.xlsx'
    headers = ['text', 'label', 'Type of PII']
    

    if os.path.exists(excel_file_path):
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)  

    
    if sheet.max_row == 1: 
        sheet.append(headers)  
    
    
    try:
        for entry in data:
            sheet.append([entry['text'], entry['label'], entry['type_of_pii']])
        workbook.save(excel_file_path)
        print("Labeled data has been appended to 'Data_Mana.xlsx'.")
    except Exception as e:
        print(f"Error writing to Excel file: {e}")

# Main function
if __name__ == "__main__":
    print("File no?")
    file_num = input()
    file_name = f"Testing_{file_num}"
    label_and_append_data(file_name)
    

